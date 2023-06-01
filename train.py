# -*- coding: utf-8 -*-
'''
@Time    : 2023/5/30 11:11
@Author  : BruanLin
@File    : train.py

'''
import torch
from openpyxl import load_workbook
from openpyxl import Workbook

from dataset.dataset import Horse2ZebraDataset
import sys
from utils import save_checkpoint, load_checkpoint
from torch.utils.data import DataLoader
from torchvision.datasets import ImageFolder
import torch.nn as nn
import torch.optim as optim
import config
from tqdm import tqdm
from torchvision.utils import save_image
from model.discriminator_model import Discriminator
from model.genorator_model import Generator


def train_fn(disc_H, disc_Z, gen_Z, gen_H, loader, opt_disc, opt_gen, l1, mse, d_scaler, g_scaler, last_row, epoch, ws):
    # 一次训练的主要过程
    loop = tqdm(loader, leave=True)
    for idx, (zebra, horse) in enumerate(loop):
        zebra = zebra.to(config.DEVICE)
        horse = horse.to(config.DEVICE)

        # 训练判定器 desc_H 和 desc_Z
        with torch.cuda.amp.autocast():  # 自动化浮点数，节省内存
            # 1.生成器：对于马
            fake_horse = gen_H(zebra)
            D_H_real = disc_H(horse)  # 鉴别真马 （接近1）
            # requires_grad = False 节省
            D_H_fake = disc_H(fake_horse.detach())  # 鉴别生出来的马 （接近0）
            # torch.ones_like函数和torch.zeros_like函数的基本功能是根据给定张量，生成与其形状相同的全1张量或全0张量
            D_H_real_loss = mse(D_H_real, torch.ones_like(D_H_real))  # 判别器的部分损失函数
            D_H_fake_loss = mse(D_H_fake, torch.zeros_like(D_H_fake))  # 判别器的部分损失函数
            D_H_LOSS = D_H_real_loss + D_H_fake_loss

            # 2.生成器：对于斑马
            fake_zebra = gen_Z(horse)
            D_Z_real = disc_Z(zebra)  # 鉴别真斑马 （接近1）
            # requires_grad = False 节省
            D_Z_fake = disc_H(fake_zebra.detach())  # 鉴别生出来斑马 （接近0）
            # torch.ones_like函数和torch.zeros_like函数的基本功能是根据给定张量，生成与其形状相同的全1张量或全0张量
            D_Z_real_loss = mse(D_Z_real, torch.ones_like(D_Z_real))  # 判别器的部分损失函数
            D_Z_fake_loss = mse(D_Z_fake, torch.zeros_like(D_Z_fake))  # 判别器的部分损失函数
            D_Z_LOSS = D_Z_real_loss + D_Z_fake_loss

            # 3.判别器总损失
            D_LOSS = (D_H_LOSS + D_Z_LOSS) / 2

            ws.cell(row=last_row, column=2, value=D_H_LOSS.item())
            ws.cell(row=last_row, column=3, value=D_Z_LOSS.item())
            ws.cell(row=last_row, column=4, value=D_LOSS.item())

        # 优化器梯度清零
        opt_disc.zero_grad()
        d_scaler.scale(D_LOSS).backward()
        d_scaler.step(opt_disc)
        d_scaler.update()

        # 训练生成器 gen_H 和 gen_Z
        with torch.cuda.amp.autocast():  # 自动化浮点数，节省内存
            D_H_fake = disc_H(fake_horse)  # 鉴别假马 从生成器的角度说，就是要欺骗判别器使得接近1
            D_Z_fake = disc_Z(fake_zebra)  # 鉴别假斑马 从生成器的角度说，就是要欺骗判别器使得接近1
            loss_G_H = mse(D_H_fake, torch.ones_like(D_H_fake))
            loss_G_Z = mse(D_Z_fake, torch.ones_like(D_Z_fake))

            # cycle loss 循环损失
            cycle_zebra = gen_Z(fake_horse)
            cycle_horse = gen_H(fake_zebra)
            cycle_zebra_loss = l1(zebra, cycle_zebra)
            cycle_horse_loss = l1(horse, cycle_horse)

            # identity 自身损失  为了提高速度
            # identity_zebra = gen_Z(zebra)
            # identity_horse = gen_H(horse)
            # identity_zebra_loss = l1(zebra, identity_zebra)
            # identity_horse_loss = l1(horse, identity_horse)

            # 全部损失
            G_loss = (
                    loss_G_Z
                    + loss_G_H
                    + cycle_zebra_loss * config.LAMBDA_CYCLE
                    + cycle_horse_loss * config.LAMBDA_CYCLE
                    # + identity_zebra_loss * LAMBDA_IDENTITY
                    # + identity_horse_loss * LAMBDA_IDENTITY
            )

            ws.cell(row=last_row, column=5, value=loss_G_Z.item())
            ws.cell(row=last_row, column=6, value=loss_G_H.item())
            ws.cell(row=last_row, column=7, value=cycle_zebra_loss.item() * config.LAMBDA_CYCLE)
            ws.cell(row=last_row, column=8, value=cycle_horse_loss.item() * config.LAMBDA_CYCLE)
            # ws.cell(row=row, column=8, value=identity_zebra_loss.item() * LAMBDA_IDENTITY)
            # ws.cell(row=row, column=9, value=identity_horse_loss.item() * LAMBDA_IDENTITY)

        opt_gen.zero_grad()
        g_scaler.scale(G_loss).backward()
        g_scaler.step(opt_gen)
        g_scaler.update()

        # 保存图像
        if idx % 200 == 0:
            save_image(fake_horse * 0.5 + 0.5,
                       f"saved_images/scenery_epoch_{epoch}_{idx}.png")
            save_image(fake_zebra * 0.5 + 0.5,
                       f"saved_images/painting_epoch_{epoch}_{idx}.png")


# 重中之重Loss过程
def main():
    # 创建或加载现有的Excel文件
    try:
        wb = load_workbook('saved_loss/loss_sum.xlsx')
    except FileNotFoundError:
        wb = Workbook()
    # 选择要写入数据的工作表（这里选择第一个工作表）
    ws = wb.active
    # 首行特殊处理
    ws.cell(row=1, column=1, value="epoch")
    ws.cell(row=1, column=2, value="D_H_LOSS")
    ws.cell(row=1, column=3, value="D_Z_LOSS")
    ws.cell(row=1, column=4, value="D_LOSS")
    ws.cell(row=1, column=5, value="loss_G_Z")
    ws.cell(row=1, column=6, value="loss_G_H")
    ws.cell(row=1, column=7, value="cycle_zebra_loss * LAMBDA_CYCLE")
    ws.cell(row=1, column=8, value="cycle_horse_loss * LAMBDA_CYCLE")
    # ws.cell(row=row, column=8, value="identity_zebra_loss * LAMBDA_IDENTITY")
    # ws.cell(row=row, column=9, value="identity_horse_loss * LAMBDA_IDENTITY")
    # 获取当前工作表的最后一行索引
    last_row = ws.max_row

    # 初始化，输入三围图像
    disc_H = Discriminator(in_channels=3).to(config.DEVICE)  # 网络作用：鉴别是不是真马
    disc_Z = Discriminator(in_channels=3).to(config.DEVICE)  # 网络作用：鉴别是不是真斑马
    gen_H = Generator(img_channels=3, num_residuals=9).to(config.DEVICE)  # 网络作用：尽可能生成出真马
    gen_Z = Generator(img_channels=3, num_residuals=9).to(config.DEVICE)  # 网络作用：尽可能生成出真斑马
    # 创建判别器的优化器
    opt_disc = optim.Adam(
        list(disc_H.parameters()) + list(disc_Z.parameters()),  # 优化的参数，这里是两个网络一起的
        lr=config.LEARNING_RATE,
        betas=(0.5, 0.999)  # 论文提出
    )
    # 创建生成器的优化器
    opt_gen = optim.Adam(
        list(gen_Z.parameters()) + list(gen_H.parameters()),  # 优化的参数，这里是两个网络一起的
        lr=config.LEARNING_RATE,
        betas=(0.5, 0.999)  # 论文提出
    )

    # 进行损失函数配置
    L1 = nn.L1Loss()
    mse = nn.MSELoss()

    # 保存数据点
    if config.LOAD_MODEL:
        load_checkpoint(
            checkpoint_file=config.CHECKPOINT_CRITIC_H, model=disc_H, optimizer=opt_disc, lr=config.LEARNING_RATE
        )
        load_checkpoint(
            checkpoint_file=config.CHECKPOINT_CRITIC_Z, model=disc_Z, optimizer=opt_disc, lr=config.LEARNING_RATE
        )
        load_checkpoint(
            checkpoint_file=config.CHECKPOINT_GEN_H, model=gen_H, optimizer=opt_gen, lr=config.LEARNING_RATE
        )
        load_checkpoint(
            checkpoint_file=config.CHECKPOINT_GEN_Z, model=gen_Z, optimizer=opt_gen, lr=config.LEARNING_RATE
        )
    # 导入数据集
    train_dataset = Horse2ZebraDataset(root_horse=config.TRAIN_DIR + "/trainA", root_zebra=config.TRAIN_DIR + "/trainB",
                                       trans_forms=config.transforms)
    loader = DataLoader(
        train_dataset,
        batch_size=config.BATCH_SIZE,
        shuffle=True,
        num_workers=config.NUM_WORKERS,
        pin_memory=True
        # 创建 DataLoader 时设置 pin_memory=True，
        # 则返回放置在页锁定内存中的批数据，使得将内存的 Tensor 数据拷贝到 GPU 显存变得更快。
    )
    # 混合精度训练，会使得效果更好,只能用于gpu
    g_scaler = torch.cuda.amp.GradScaler()
    d_scaler = torch.cuda.amp.GradScaler()
    # 开始轮次训练
    for epoch in range(config.NUM_EPOCHS):
        # 训练主要过程
        train_fn(disc_H, disc_Z, gen_Z, gen_H, loader, opt_disc, opt_gen, L1, mse, d_scaler, g_scaler,
                 epoch + 1, epoch + 1, ws)

        # 保存数据
        wb.save('saved_loss/loss_sum.xlsx')
        if config.SAVE_MODEL:
            save_checkpoint(gen_H, opt_gen, filename=config.CHECKPOINT_GEN_H)
            save_checkpoint(gen_Z, opt_gen, filename=config.CHECKPOINT_GEN_Z)
            save_checkpoint(disc_H, opt_disc, filename=config.CHECKPOINT_CRITIC_H)
            save_checkpoint(disc_Z, opt_disc, filename=config.CHECKPOINT_CRITIC_Z)


if __name__ == "__main__":
    main()
